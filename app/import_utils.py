import csv
import io
import re
from pathlib import Path

from openpyxl import load_workbook


EMAIL_PATTERN = re.compile(
    r"^[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}$",
    re.IGNORECASE,
)


def normalize_email(value: object) -> str:
    return str(value or "").strip().lower()


def is_valid_email(value: str) -> bool:
    return bool(EMAIL_PATTERN.fullmatch(value))


def extract_from_txt(content: bytes) -> list[str]:
    text = content.decode("utf-8-sig", errors="ignore")

    values = re.split(r"[\s,;|]+", text)
    return [normalize_email(value) for value in values if value.strip()]


def extract_from_csv(content: bytes) -> list[str]:
    text = content.decode("utf-8-sig", errors="ignore")

    sample = text[:4096]

    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel

    reader = csv.reader(io.StringIO(text), dialect=dialect)

    values: list[str] = []

    for row in reader:
        for cell in row:
            email = normalize_email(cell)

            if "@" in email:
                values.append(email)

    return values


def extract_from_xlsx(content: bytes) -> list[str]:
    workbook = load_workbook(
        filename=io.BytesIO(content),
        read_only=True,
        data_only=True,
    )

    values: list[str] = []

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows(values_only=True):
            for cell in row:
                email = normalize_email(cell)

                if "@" in email:
                    values.append(email)

    workbook.close()
    return values


def extract_emails(filename: str, content: bytes) -> list[str]:
    extension = Path(filename).suffix.lower()

    if extension == ".txt":
        return extract_from_txt(content)

    if extension == ".csv":
        return extract_from_csv(content)

    if extension == ".xlsx":
        return extract_from_xlsx(content)

    raise ValueError(
        "Nepodprta datoteka. Uporabi CSV, XLSX ali TXT."
    )